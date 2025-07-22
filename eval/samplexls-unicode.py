import pandas as pd
import openpyxl
from openpyxl import load_workbook
import xlrd

# Method 1: Using pandas read_excel (Recommended)
def read_excel_with_unicode_pandas(file_path, sheet_name=None):
    """
    Read Excel file with Unicode characters using pandas
    Pandas handles Unicode automatically for Excel files
    """
    try:
        # Read Excel file - pandas handles Unicode automatically
        if sheet_name:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
        else:
            df = pd.read_excel(file_path)
        
        print(f"✓ Successfully read Excel file")
        print(f"Shape: {df.shape}")
        print(f"Columns: {df.columns.tolist()}")
        
        return df
    
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None

# Method 2: Read specific sheet by name or index
def read_excel_specific_sheet(file_path, sheet_identifier=0):
    """
    Read specific Excel sheet with Unicode support
    sheet_identifier can be sheet name (string) or index (int)
    """
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_identifier)
        print(f"✓ Successfully read sheet: {sheet_identifier}")
        return df
    except Exception as e:
        print(f"Error reading sheet {sheet_identifier}: {e}")
        return None

# Method 3: Read all sheets at once
def read_all_excel_sheets(file_path):
    """
    Read all sheets from Excel file into a dictionary
    """
    try:
        # Read all sheets
        all_sheets = pd.read_excel(file_path, sheet_name=None)
        
        print(f"✓ Found {len(all_sheets)} sheets:")
        for sheet_name in all_sheets.keys():
            print(f"  - {sheet_name}: {all_sheets[sheet_name].shape}")
        
        return all_sheets
    
    except Exception as e:
        print(f"Error reading all sheets: {e}")
        return None

# Method 4: Using openpyxl directly for more control
def read_excel_openpyxl(file_path, sheet_name=None):
    """
    Read Excel using openpyxl for more control over Unicode handling
    """
    try:
        # Load workbook
        workbook = load_workbook(file_path, data_only=True)
        
        # Get sheet names
        sheet_names = workbook.sheetnames
        print(f"Available sheets: {sheet_names}")
        
        # Select sheet
        if sheet_name:
            if sheet_name in sheet_names:
                worksheet = workbook[sheet_name]
            else:
                print(f"Sheet '{sheet_name}' not found. Using first sheet.")
                worksheet = workbook.active
        else:
            worksheet = workbook.active
        
        # Extract data
        data = []
        headers = []
        
        # Get headers from first row
        for cell in worksheet[1]:
            headers.append(cell.value)
        
        # Get data from remaining rows
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            data.append(list(row))
        
        # Create DataFrame
        df = pd.DataFrame(data, columns=headers)
        
        print(f"✓ Successfully read using openpyxl")
        print(f"Shape: {df.shape}")
        
        workbook.close()
        return df
    
    except Exception as e:
        print(f"Error using openpyxl: {e}")
        return None

# Method 5: Handle Excel files with different engines
def read_excel_with_engine(file_path, engine='openpyxl'):
    """
    Read Excel with specific engine
    Available engines: 'openpyxl', 'xlrd', 'odf', 'pyxlsb'
    """
    engines_to_try = [engine, 'openpyxl', 'xlrd']
    
    for eng in engines_to_try:
        try:
            print(f"Trying engine: {eng}")
            df = pd.read_excel(file_path, engine=eng)
            print(f"✓ Successfully read with {eng} engine")
            return df
        except Exception as e:
            print(f"✗ Failed with {eng}: {e}")
            continue
    
    print("All engines failed")
    return None

# Method 6: Advanced reading with parameters
def read_excel_advanced(file_path, **kwargs):
    """
    Advanced Excel reading with custom parameters
    """
    default_params = {
        'sheet_name': 0,
        'header': 0,
        'skiprows': None,
        'nrows': None,
        'usecols': None,
        'engine': 'openpyxl'
    }
    
    # Update with user parameters
    params = {**default_params, **kwargs}
    
    try:
        df = pd.read_excel(file_path, **params)
        print(f"✓ Successfully read Excel with advanced parameters")
        print(f"Shape: {df.shape}")
        return df
    except Exception as e:
        print(f"Error with advanced reading: {e}")
        return None

# Method 7: Handle Excel files with merged cells and formatting
def read_excel_handle_merged_cells(file_path, sheet_name=None):
    """
    Read Excel file and handle merged cells properly
    """
    try:
        workbook = load_workbook(file_path)
        
        if sheet_name:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.active
        
        # Unmerge all merged cells first
        merged_ranges = list(worksheet.merged_cells.ranges)
        for merged_range in merged_ranges:
            worksheet.unmerge_cells(str(merged_range))
        
        # Convert to pandas DataFrame
        data = worksheet.values
        columns = next(data)[0:]
        df = pd.DataFrame(data, columns=columns)
        
        workbook.close()
        return df
    
    except Exception as e:
        print(f"Error handling merged cells: {e}")
        return None

# Utility functions
def inspect_excel_file(file_path):
    """
    Inspect Excel file structure
    """
    try:
        # Get basic info
        xl_file = pd.ExcelFile(file_path)
        print(f"Excel file info:")
        print(f"  Engine: {xl_file.engine}")
        print(f"  Sheet names: {xl_file.sheet_names}")
        
        # Check each sheet
        for sheet_name in xl_file.sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=0)
            print(f"  Sheet '{sheet_name}': {len(df.columns)} columns")
            print(f"    Columns: {df.columns.tolist()}")
        
        xl_file.close()
        
    except Exception as e:
        print(f"Error inspecting file: {e}")

def clean_unicode_dataframe(df):
    """
    Clean Unicode characters in DataFrame
    """
    # Clean column names
    df.columns = df.columns.astype(str)  # Convert to string
    df.columns = df.columns.str.replace('\ufeff', '')  # Remove BOM
    df.columns = df.columns.str.strip()  # Remove whitespace
    
    # Clean string columns
    string_columns = df.select_dtypes(include=['object']).columns
    for col in string_columns:
        df[col] = df[col].astype(str).str.replace('\ufeff', '')
        df[col] = df[col].str.strip()
    
    return df

# Example usage
def main():
    file_path = '../textsamples.xlsx'  # Replace with your file path
    
    print("=== Basic Excel Reading ===")
    df = read_excel_with_unicode_pandas(file_path)
    if df is not None:
        print("First few rows:")
        print(df.head())
        
        # Check for Unicode in columns
        print("\nColumn names with potential Unicode:")
        for col in df.columns:
            if any(ord(char) > 127 for char in str(col)):
                print(f"  Unicode in column: {col}")
    
    print("\n=== Inspect File Structure ===")
    inspect_excel_file(file_path)
    
    print("\n=== Read All Sheets ===")
    all_sheets = read_all_excel_sheets(file_path)
    
    print("\n=== Clean Unicode Characters ===")
    if df is not None:
        cleaned_df = clean_unicode_dataframe(df.copy())
        print("Cleaned column names:", cleaned_df.columns.tolist())

# Quick usage examples
def quick_examples():
    """
    Quick examples for common scenarios
    """
    file_path = '../samplexls-unicode.xlsx'
    
    # Basic reading
    df = pd.read_excel(file_path)
    
    # Read specific sheet
    df = pd.read_excel(file_path, sheet_name='Sheet1')
    
    # Read with specific columns
    df = pd.read_excel(file_path, usecols=['col1', 'col2', 'col3'])
    
    # Skip rows
    df = pd.read_excel(file_path, skiprows=2)
    
    # Read only first N rows
    df = pd.read_excel(file_path, nrows=100)

if __name__ == "__main__":
    # Replace with your file path
    main()
    
    # Simple usage:
    # df = pd.read_excel('your_file.xlsx')
    # print(df.head())
    pass